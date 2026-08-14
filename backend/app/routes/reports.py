from io import BytesIO

from flask import Blueprint, g, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..services.report_service import customer_report, monthly_report, pending_report
from ..utils import ValidationError, api_error, parse_int_in_range
from .common import require_auth

reports_bp = Blueprint("reports", __name__)


def report_from_request():
    report_type = request.args.get("type", "monthly")
    if report_type not in {"monthly", "customers", "pending"}:
        raise ValidationError("Report type must be monthly, customers, or pending.", "type")
    if report_type == "customers":
        return report_type, customer_report(g.agent.id)
    from datetime import date
    today = date.today()
    month = parse_int_in_range(request.args.get("month", today.month), "month", 1, 12)
    year = parse_int_in_range(request.args.get("year", today.year), "year", 2000, 2200)
    report = monthly_report(g.agent.id, month, year) if report_type == "monthly" else pending_report(g.agent.id, month, year)
    return report_type, report


def report_excel(report):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RD Report"
    sheet.append([report["title"]])
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="155E75")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report["columns"]))
    sheet.append([])
    labels = [column.replace("_", " ").title() for column in report["columns"]]
    sheet.append(labels)
    for cell in sheet[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
    for row in report["rows"]:
        sheet.append([row.get(column, "") for column in report["columns"]])
    sheet.append([])
    for key, value in report["summary"].items():
        sheet.append([key.replace("_", " ").title(), value])
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 14), 54)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def report_pdf(report):
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=22, leftMargin=22, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    labels = [column.replace("_", " ").title() for column in report["columns"]]
    data = [labels] + [[str(row.get(column, ""))[:70] for column in report["columns"]] for row in report["rows"]]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDFA")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    summary = ", ".join(f"{key.replace('_', ' ').title()}: {value}" for key, value in report["summary"].items())
    document.build([Paragraph(report["title"], styles["Title"]), Spacer(1, 8), Paragraph(summary, styles["Normal"]), Spacer(1, 10), table])
    output.seek(0)
    return output


@reports_bp.get("/")
@require_auth
def get_report():
    try:
        _, report = report_from_request()
    except ValidationError as exc:
        return api_error(exc.message, 400, exc.field)
    return jsonify(report)


@reports_bp.get("/export")
@require_auth
def export_report():
    try:
        report_type, report = report_from_request()
    except ValidationError as exc:
        return api_error(exc.message, 400, exc.field)
    export_format = request.args.get("format", "xlsx")
    if export_format == "xlsx":
        return send_file(report_excel(report), as_attachment=True, download_name=f"rd-{report_type}-report.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if export_format == "pdf":
        return send_file(report_pdf(report), as_attachment=True, download_name=f"rd-{report_type}-report.pdf", mimetype="application/pdf")
    return api_error("Export format must be xlsx or pdf.", 400, "format")
